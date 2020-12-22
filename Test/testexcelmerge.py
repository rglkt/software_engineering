import sys
import os
sys.path.append(os.path.split(sys.path[0])[0])
from fun.mergeExcel import mergeExcel
from fun.mytool import _enumerate_filename
import openpyxl

def testExcelMerge(directory,row,col):
    test_row = row-1
    file_list = _enumerate_filename(directory)
    for file in file_list:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        test_row+=ws.max_row+1-row
    mergeExcel(directory,row,col)
    wb = openpyxl.load_workbook(os.path.join(directory,'mergeExcel.xlsx'))
    ws = wb.active
    if ws.max_row == test_row:
        print("Success!")
    else :
        print("Error!")
    os.remove(os.path.join(directory,'mergeExcel.xlsx'))

if __name__ == '__main__':
    directory = "./TestData/Excel/"
    row = 2
    col = 1
    testExcelMerge(directory,row,col)