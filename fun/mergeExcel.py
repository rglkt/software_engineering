                        
import openpyxl
import os 
from .mytool import _enumerate_filename
def mergeExcel(path,row=2,col=1):
    file_list = _enumerate_filename(path)
    wb = openpyxl.load_workbook(file_list[0])
    ws = wb.active 
    for file_name in file_list[1:]:
        wb_new = openpyxl.load_workbook(file_name)
        ws_new = wb_new.active
        maxrow_temp = ws.max_row
        for i in range(row,ws_new.max_row+1):
            for j in range(col,ws_new.max_column+1):
                ws.cell(maxrow_temp+i-row+1,j-col+1,ws_new.cell(i,j).value)
    wb.save(os.path.join(path, "mergeExcel.xlsx"))

